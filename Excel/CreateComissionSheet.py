import openpyxl

planilha = openpyxl.Workbook()

print(planilha.sheetnames)

planilha.create_sheet('Comissão')

comissao_page = planilha['Comissão']
comissao_page.append(['Vendedor', ' Data', 'Comissão'])
comissao_page.append(['Alice', ' 2023-01-01', 150.90])
comissao_page.append(['Arthur', ' 2023-01-03', 100.80])
comissao_page.append(['Matheus', ' 2023-01-04', 90.90])

planilha.save('Planilha_de_Comissoes.xlsx')
