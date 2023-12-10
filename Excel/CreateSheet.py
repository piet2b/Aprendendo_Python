import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Frutas')

frutas_page = book['Frutas']
frutas_page.append(['Frutas', ' Quantidade', 'Preço'])
frutas_page.append(['Banana', ' 5', 'R$3,90'])
frutas_page.append(['Maça', ' 6', 'R$8,20'])
frutas_page.append(['Abacaxi', ' 2', 'R$6,80'])
frutas_page.append(['Mamão', ' 1', 'R$9,90'])

book.save('Planilha de compras.xlsx')
