import openpyxl

book = openpyxl.load_workbook('Planilha de compras.xlsx')

frutas_page = book['Frutas']

for rows in frutas_page.iter_rows(min_row=1, max_row=4):
    print(f'{rows[0].value},{rows[1].value}, {rows[2].value}')

for rows in frutas_page.iter_rows(min_row=1, max_row=4):
    for cell in rows:
        if cell.value == 'Banana':
            cell.value = 'Fruta 1'

book.save('Planilha de compras v2.xlsx')
